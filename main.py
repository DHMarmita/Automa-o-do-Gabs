import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from concurrent.futures import ThreadPoolExecutor, as_completed
from buscar_google import buscar_informacoes_econodata, buscar_informacoes_google

# Função para normalizar string para UPPERCASE e sem espaços em branco no final


def normalizar_valor(valor):
    if isinstance(valor, str):
        # Remove espaços em branco no início e fim, e converte para UPPERCASE
        return valor.strip().upper()
    return 'N/A'  # Retorna 'N/A' se o valor não for uma string (caso de NaN)

# Função para buscar informações utilizando threads


def buscar_informacoes_thread(registro, resultados_anteriores):
    cliente = registro['Cliente']
    estado = registro['Estado']
    cidade = registro['Cidade']
    chave = (cliente, estado, cidade)

    if chave in resultados_anteriores:
        retorno_cnpj, retorno_endereco = resultados_anteriores[chave]
    else:
        # Realiza a busca de informações do cnpj.biz
        retorno_cnpj = buscar_informacoes_google(cliente, estado, cidade)
        # Realiza a busca de informações do econodata.com.br
        retorno_endereco = buscar_informacoes_econodata(retorno_cnpj['CNPJ'])
        # Armazena os resultados para futuras referências
        resultados_anteriores[chave] = (retorno_cnpj, retorno_endereco)

    # Atualiza o registro com as informações encontradas
    registro.update(retorno_cnpj)
    registro.update(retorno_endereco)


# Caminho para o arquivo Excel
caminho_arquivo = 'arquivo.xlsx'

# Carregando o arquivo Excel
xls = pd.ExcelFile(caminho_arquivo)

# Lista para armazenar os registros de todas as planilhas
registros = []

# Dicionário para armazenar os resultados anteriores
resultados_anteriores = {}

# Iterando sobre todas as planilhas no arquivo Excel
for nome_planilha in xls.sheet_names:
    # Carregando a planilha atual
    df = pd.read_excel(xls, sheet_name=nome_planilha)

    # Iterando pelas linhas da planilha atual
    for indice, linha in df.iterrows():
        # Lendo os campos e aplicando a normalização
        equipamento = normalizar_valor(linha.get('Equipamento', ''))
        cliente = normalizar_valor(linha.get('Cliente', ''))
        estado = normalizar_valor(linha.get('Estado', ''))
        cidade = normalizar_valor(linha.get('Cidade', ''))
        marca_icp = normalizar_valor(linha.get('Marca ICP', ''))
        modelo = normalizar_valor(linha.get('Modelo', ''))

        # Criando um dicionário para representar a linha atual
        registro = {
            'Equipamento': equipamento,
            'Cliente': cliente,
            'Estado': estado,
            'Cidade': cidade,
            'Marca ICP': marca_icp,
            'Modelo': modelo,
            'Planilha': nome_planilha  # Adicionando o nome da planilha
        }

        chave = (cliente, estado, cidade)
        if chave in resultados_anteriores:
            # Preenche o registro com as informações anteriores
            retorno_cnpj, retorno_endereco = resultados_anteriores[chave]
            registro.update(retorno_cnpj)
            registro.update(retorno_endereco)
            # Atualiza o DataFrame com as informações
            df.at[indice, 'CNPJ'] = retorno_cnpj.get('CNPJ', 'N/A')
            df.at[indice, 'CEP'] = retorno_endereco.get('CEP', 'N/A')
            df.at[indice, 'Logradouro'] = retorno_endereco.get('Logradouro', 'N/A')
            df.at[indice, 'Bairro'] = retorno_endereco.get('Bairro', 'N/A')
            df.at[indice, 'Nome Fantasia'] = retorno_endereco.get('Nome Fantasia', 'N/A')
        else:
            # Adicionando o registro à lista de registros
            registros.append(registro)

# Definindo o número máximo de threads
max_threads = 10

# Criando um ThreadPoolExecutor para limitar a quantidade de threads
with ThreadPoolExecutor(max_workers=max_threads) as executor:
    # Criando uma lista para armazenar os futuros
    futures = [executor.submit(
        buscar_informacoes_thread, registro, resultados_anteriores) for registro in registros]

    # Aguardando todas as threads terminarem
    for future in as_completed(futures):
        future.result()

# Convertendo a lista de registros em um DataFrame do pandas
df_registros = pd.DataFrame(registros)

# Escrevendo o DataFrame em um arquivo Excel temporário
caminho_saida_excel = 'registros_temp.xlsx'
df_registros.to_excel(caminho_saida_excel, index=False)

# Abrindo o arquivo Excel com openpyxl para aplicar a formatação condicional
wb = load_workbook(caminho_saida_excel)
ws = wb.active

# Definindo o estilo de preenchimento para células com "N/A"
fill_na = PatternFill(start_color='FFCCCC',
                      end_color='FFCCCC', fill_type='solid')

# Aplicando a formatação condicional
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        if cell.value == "N/A" and cell.column not in ('Marca ICP', 'Modelo'):
            cell.fill = fill_na

# Salvando o arquivo Excel final
caminho_saida_excel_final = 'registros.xlsx'
wb.save(caminho_saida_excel_final)

print(f"Arquivo Excel gerado com sucesso: {caminho_saida_excel_final}  Total: {len(registros)}")
